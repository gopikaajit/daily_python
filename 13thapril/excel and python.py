import openpyxl 


print("Hello world!")

wb = openpyxl.load_workbook('excel and python.xlsx')
print(wb)
ws = wb.active
print(ws)
print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
print('The value in cell A1 is: '+ws['A1'].value)
values = [ws.cell(row=1,column=i).value for i in range(1,ws.max_column+1)]
print(values)
values = [ws.cell(row=2,column=i).value for i in range(1,ws.max_column+1)]
print(values)
ws.cell(row=3, column=1, value = 'Total')


values = [ws.cell(row=3,column=i).value for i in range(1,2)]
print(values)

#finding sum using formula
ws.cell(row=3, column=2).value = "=SUM(B2:M2)" 
wb.save("excel and python2.xlsx")

